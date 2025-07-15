import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Load the Excel file
file_path = r"C:\Users\tin.xu\Desktop\Report\on_water_report.xlsx"
xls = pd.ExcelFile(file_path)

# Dictionary to store row counts for each sheet
sheet_counts = {}

# Initialize set to track unique container numbers
unique_containers = set()

# Iterate through all sheets and count the number of valid rows
for sheet in xls.sheet_names:
    try:
        df_sheet = pd.read_excel(xls, sheet_name=sheet)

        if "Container#" in df_sheet.columns:
            valid_containers = df_sheet["Container#"].dropna().unique()
            unique_containers.update(valid_containers)  # Add unique containers

            # Store sheet-wise count
            sheet_counts[sheet] = len(valid_containers)

    except Exception as e:
        print(f"⚠️ Error processing sheet '{sheet}': {e}")
        pass  # Ignore errors (e.g., empty sheets)

# Count the unique container numbers
unique_container_count = len(unique_containers)

# Creating DataFrame for summary
df_summary = pd.DataFrame(sheet_counts.items(), columns=["WM", "Qty of Ctns"])

# Add a numbering column, starting from A3
df_summary.insert(0, "No.", range(1, len(df_summary) + 1))

# Add a total row at the bottom
total_value = df_summary["Qty of Ctns"].sum()
df_summary.loc[len(df_summary)] = ["", "Total On Water Ctns", total_value]

# Save DataFrame to an Excel file
summary_path = r"C:\Users\tin.xu\Desktop\Report\on_water_summary.xlsx"
df_summary.to_excel(summary_path, index=False, engine="openpyxl")

# Load the saved Excel file to apply formatting
wb = load_workbook(summary_path)
ws = wb.active

# Add title
ws.insert_rows(1)
ws["A1"] = "On Water Summary"
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df_summary.columns))

# Define styles
grey_blue_fill = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid")
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
black_font = Font(size=12, bold=True, color="000000")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Style title
title_cell = ws["A1"]
title_cell.font = Font(size=14, bold=True, color="000000")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
title_cell.fill = grey_blue_fill

# Style headers
for cell in ws[2]:
    cell.fill = orange_fill
    cell.font = black_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Adjust column widths dynamically
for col_idx, col_cells in enumerate(ws.columns, start=1):
    max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells) + 2
    adjusted_width = min(max_length, 20)  # Cap max width to 20
    ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

# Apply borders and formatting
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
    for cell in row:
        if cell.value not in [None, " ", ""]:
            cell.border = thin_border
        cell.font = Font(color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")

# Remove background color for total row
last_row_index = ws.max_row
for cell in ws[last_row_index]:
    cell.fill = PatternFill(fill_type="none")
    cell.font = Font(bold=True, color="000000")
    cell.border = thin_border

# Save the modified workbook
wb.save(summary_path)

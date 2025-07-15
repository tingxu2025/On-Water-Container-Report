from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill

# Define the path to your Excel file
file_path = r"C:\Users\tin.xu\Desktop\Report\on_water_report.xlsx"
save_path = r"C:\Users\tin.xu\Desktop\Report\chinese_on_water_report.xlsx"
# Load the workbook
wb = load_workbook(file_path)

# Define the border style
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Define fills
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
grey_blue_fill = PatternFill(start_color='B0C4DE', end_color='B0C4DE', fill_type='solid')  # Grey blue for column J (from row 3)

# Define fonts
first_row_font = Font(name="Calibri", size=14, bold=False)  # First row font
chinese_row_font = Font(name="Calibri", size=14, bold=False)  # Second row font (Chinese headers)
default_font = Font(name="Arial", size=10)  # Default font for all other cells
bold_font = Font(bold=True)  # Bold font for specific formatting

# Define alignment (center alignment for all cells)
center_alignment = Alignment(horizontal='center', vertical='center')

# Define the second row headers (Chinese headers)
chinese_headers = ["编号", "柜号", "数量", "预计开船日期", "预计到达日期", "收柜日期", "出货点", "备注", "船运公司", "目前状况", "发票金额", "已出发票日期"]

# Process each sheet in the workbook
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Check and delete the "Destination" column if it exists
    header_values = [cell.value for cell in ws[1]]
    if "Destination" in header_values:
        destination_col_index = header_values.index("Destination") + 1  # Get the column index
        ws.delete_cols(destination_col_index)

    # Insert Chinese headers as the second row and shift existing rows down
    ws.insert_rows(2)
    for index, header in enumerate(chinese_headers, start=1):
        ws.cell(row=2, column=index).value = header
        ws.cell(row=2, column=index).font = chinese_row_font
        ws.cell(row=2, column=index).alignment = center_alignment
        ws.cell(row=2, column=index).border = thin_border

    # Apply formatting to all cells
    for row in ws.iter_rows(min_row=3, max_col=ws.max_column):  # Start formatting from row 3
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_alignment

            # Apply specific formatting based on cell position
            if cell.row == 2:
                cell.font = chinese_row_font
                cell.fill = yellow_fill
            elif cell.column_letter == 'J' and cell.row >= 3:  # Grey blue fill for column J (from row 3)
                cell.fill = grey_blue_fill
                cell.font = bold_font
            else:
                cell.font = default_font

    # Explicitly reapply formatting for the first row
    for cell in ws[1]:
        cell.font = first_row_font
        cell.fill = yellow_fill
        cell.border = thin_border
        cell.alignment = center_alignment

    # Adjust column widths based on maximum content length
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter  # Get column letter (e.g., 'A', 'B', etc.)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        adjusted_width = (max_length + 5) * 1.2  # Add some padding for better readability
        ws.column_dimensions[column_letter].width = adjusted_width

# Save the workbook with the updated headers, formatting, and adjusted column widths
wb.save(save_path)
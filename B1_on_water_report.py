import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

# Define file paths
file_path = r"C:\Users\tin.xu\Desktop\Report\WAHHUNG-Sterling.xlsx"  # Update the path
save_path = r"C:\Users\tin.xu\Desktop\Report\on_water_report.xlsx"

# Load the Excel file
df_sterling = pd.ExcelFile(file_path)

# Load the 2024 sheet
df_sterling_2024 = df_sterling.parse("2024")

# Ensure the correct column names
delivery_date_col = "Delivery Date"
destination_col = "Destination"
eta_col = "ETA"
etd_col = "ETD"

# Filter rows where "Delivery Date" is empty (NaN)
empty_delivery_dates = df_sterling_2024[df_sterling_2024[delivery_date_col].isna()].copy()

# Convert ETA and ETD to date format (MM/DD/YYYY)
for date_col in [eta_col, etd_col]:
    if date_col in empty_delivery_dates.columns:
        empty_delivery_dates[date_col] = pd.to_datetime(empty_delivery_dates[date_col], errors='coerce').dt.strftime('%m/%d/%Y')

# Add missing columns with default values
empty_delivery_dates["Status"] = "ON WATER"  # Set Status as "on water"
empty_delivery_dates["Invoice Amount"] = np.nan  # Placeholder for missing Invoice Amount
empty_delivery_dates["Issued Invoice Date"] = pd.NaT  # Placeholder for missing Issued Invoice Date

# Convert "Invoice Amount" to numeric to prevent date misinterpretation
empty_delivery_dates["Invoice Amount"] = pd.to_numeric(empty_delivery_dates["Invoice Amount"], errors='coerce')

# Convert "Issued Invoice Date" to datetime, keeping NaT values
empty_delivery_dates["Issued Invoice Date"] = pd.to_datetime(empty_delivery_dates["Issued Invoice Date"], errors='coerce')

# Reorder columns as specified
ordered_columns = [
    "Ref#", "Container#", "Qty", "ETD", "ETA", "Delivery Date", "Origin",
    "Note", "Carrier", "Status", "Invoice Amount", "Issued Invoice Date", "Destination"
]

# Ensure column existence before reordering
available_columns = [col for col in ordered_columns if col in empty_delivery_dates.columns]
empty_delivery_dates = empty_delivery_dates[available_columns]

# Define the desired sheet order
sheet_order = ["ELF", "ELO", "ELG", "DAL", "HOU", "SAC", "NY", "DC"]

# Create an Excel writer object to store multiple sheets
with pd.ExcelWriter(save_path, engine='xlsxwriter') as writer:
    for sheet_name in sheet_order:
        # Filter data for the specific destination
        df_filtered = empty_delivery_dates[empty_delivery_dates[destination_col] == sheet_name]

        # Ensure a sheet is created even if empty
        if df_filtered.empty:
            df_filtered = pd.DataFrame(columns=available_columns)  # Create an empty DataFrame with the correct structure

        # Write to Excel sheet
        df_filtered.to_excel(writer, sheet_name=sheet_name, index=False)

print(f"file saved to: {save_path}")

# Load the workbook and adjust column widths and alignment
wb = load_workbook(save_path)

extra_padding = 5  # Increase padding for extra space
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow background

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Auto-adjust column width with extra padding
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get column letter

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = max_length + extra_padding  # Add extra padding for better visibility
        ws.column_dimensions[col_letter].width = adjusted_width

    # Apply yellow background color to the header row (Row 1)
    for cell in ws[1]:  # First row is the header row
        cell.fill = yellow_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Apply correct formatting and center alignment for all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")  # Center the text

            if cell.column_letter == "J":  # Invoice Amount column
                cell.number_format = '0.00'  # Two decimal places

            if cell.column_letter == "K":  # Issued Invoice Date column
                if isinstance(cell.value, str):
                    try:
                        cell.value = pd.to_datetime(cell.value).strftime('%m/%d/%Y')  # Format as MM/DD/YYYY
                    except:
                        pass  # Ignore non-date values

# Save the modified workbook
wb.save(save_path)